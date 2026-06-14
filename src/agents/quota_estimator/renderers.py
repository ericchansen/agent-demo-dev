"""Artifact renderers for quota estimation reports."""

from __future__ import annotations

import html
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")

from matplotlib import pyplot as plt  # noqa: E402
from matplotlib.backends.backend_pdf import PdfPages  # noqa: E402
from matplotlib.ticker import FuncFormatter  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from src.agents.quota_estimator.models import QuotaEstimate

_HEADER_FILL = PatternFill("solid", fgColor="0078D4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(size=16, bold=True)


def render_quota_xlsx(estimate: QuotaEstimate, output_path: str | Path) -> str:
    """Render an Excel workbook containing summary, recommendation, and source tabs."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet["A1"] = f"Quota Estimate - {estimate.customer_name}"
    summary_sheet["A1"].font = _TITLE_FONT
    summary_rows = [
        ("Generated at", estimate.generated_at.isoformat(timespec="seconds")),
        ("Trailing revenue", estimate.trailing_revenue_total),
        ("Recommended quota", estimate.recommended_quota_total),
        ("Overall growth rate", estimate.overall_growth_rate),
        ("WorkIQ engagement", estimate.workiq_activity.engagement_score),
        ("WorkIQ activity count", estimate.workiq_activity.activity_count),
    ]
    for row_index, row in enumerate(summary_rows, start=3):
        summary_sheet.cell(row=row_index, column=1, value=row[0])
        summary_sheet.cell(row=row_index, column=2, value=row[1])
    summary_sheet["B4"].number_format = "$#,##0"
    summary_sheet["B5"].number_format = "$#,##0"
    summary_sheet["B6"].number_format = "0.0%"

    recommendation_sheet = workbook.create_sheet("Recommendations")
    _append_header(
        recommendation_sheet,
        [
            "Territory",
            "Category",
            "Trailing Revenue",
            "Quantity",
            "Historical Growth",
            "Market Adj.",
            "Engagement Adj.",
            "Recommended Growth",
            "Recommended Quota",
            "Rationale",
        ],
    )
    for recommendation in estimate.recommendations:
        recommendation_sheet.append(
            [
                recommendation.territory,
                recommendation.category,
                recommendation.trailing_revenue,
                recommendation.trailing_quantity,
                recommendation.historical_growth_rate,
                recommendation.market_adjustment,
                recommendation.engagement_adjustment,
                recommendation.recommended_growth_rate,
                recommendation.recommended_quota,
                recommendation.rationale,
            ]
        )
    for row in recommendation_sheet.iter_rows(min_row=2, min_col=3, max_col=9):
        row[0].number_format = "$#,##0"
        row[1].number_format = "#,##0"
        for cell in row[2:6]:
            cell.number_format = "0.0%"
        row[6].number_format = "$#,##0"

    detail_sheet = workbook.create_sheet("Sales Detail")
    _append_header(detail_sheet, ["Territory", "Category", "Order Date", "Revenue", "Quantity"])
    for sales_row in estimate.sales_rows:
        detail_sheet.append(
            [
                sales_row.territory,
                sales_row.category,
                sales_row.order_date.isoformat(),
                sales_row.revenue,
                sales_row.quantity,
            ]
        )
    for row in detail_sheet.iter_rows(min_row=2, min_col=4, max_col=5):
        row[0].number_format = "$#,##0"
        row[1].number_format = "#,##0"

    methodology_sheet = workbook.create_sheet("Methodology")
    methodology_sheet["A1"] = "Methodology"
    methodology_sheet["A1"].font = _TITLE_FONT
    methodology_sheet["A2"] = estimate.methodology
    methodology_sheet["A4"] = "Market Context"
    methodology_sheet["A4"].font = Font(bold=True)
    methodology_sheet["A5"] = estimate.research_context.summary
    methodology_sheet["A7"] = "Citations"
    methodology_sheet["A7"].font = Font(bold=True)
    for row_index, citation in enumerate(estimate.citations, start=8):
        methodology_sheet.cell(row=row_index, column=1, value=citation)

    for sheet in workbook.worksheets:
        _autosize_columns(sheet)

    workbook.save(path)
    return str(path.resolve())


def render_quota_html(estimate: QuotaEstimate, output_path: str | Path) -> str:
    """Render a complete HTML quota report."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    recommendation_rows = "\n".join(
        "<tr>"
        f"<td>{html.escape(item.territory)}</td>"
        f"<td>{html.escape(item.category)}</td>"
        f"<td>{_currency(item.trailing_revenue)}</td>"
        f"<td>{_percent(item.recommended_growth_rate)}</td>"
        f"<td>{_currency(item.recommended_quota)}</td>"
        f"<td>{html.escape(item.rationale)}</td>"
        "</tr>"
        for item in estimate.recommendations
    )
    citations = "\n".join(f"<li>{html.escape(citation)}</li>" for citation in estimate.citations)

    document = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Quota Estimate - {html.escape(estimate.customer_name)}</title>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin: 2rem; color: #1f1f1f; }}
    table {{ border-collapse: collapse; width: 100%; margin: 1rem 0; }}
    th, td {{ border: 1px solid #d0d7de; padding: 0.5rem; text-align: left; }}
    th {{ background: #0078d4; color: white; }}
    .metric {{ display: inline-block; margin-right: 2rem; font-size: 1.1rem; }}
    .metric strong {{ display: block; color: #0078d4; }}
  </style>
</head>
<body>
  <h1>Quota Estimate - {html.escape(estimate.customer_name)}</h1>
  <p>Generated {html.escape(estimate.generated_at.isoformat(timespec="seconds"))}</p>
  <section>
    <div class="metric"><strong>{_currency(estimate.trailing_revenue_total)}</strong>Trailing revenue</div>
    <div class="metric"><strong>{_currency(estimate.recommended_quota_total)}</strong>Recommended quota</div>
    <div class="metric"><strong>{_percent(estimate.overall_growth_rate)}</strong>Overall growth</div>
  </section>
  <h2>Recommendations</h2>
  <table>
    <thead>
      <tr>
        <th>Territory</th><th>Category</th><th>Trailing Revenue</th>
        <th>Growth</th><th>Quota</th><th>Rationale</th>
      </tr>
    </thead>
    <tbody>{recommendation_rows}</tbody>
  </table>
  <h2>Market and WorkIQ Context</h2>
  <p>{html.escape(estimate.research_context.summary)}</p>
  <p>WorkIQ engagement: {html.escape(estimate.workiq_activity.engagement_score)}
     ({estimate.workiq_activity.activity_count} recent activities from
     {html.escape(estimate.workiq_activity.source)}).</p>
  <h2>Methodology</h2>
  <p>{html.escape(estimate.methodology)}</p>
  <h2>Citations</h2>
  <ol>{citations}</ol>
</body>
</html>
"""
    path.write_text(document, encoding="utf-8")
    return str(path.resolve())


def render_quota_pdf(estimate: QuotaEstimate, output_path: str | Path) -> str:
    """Render a PDF report using matplotlib's built-in PDF backend."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    with PdfPages(path) as pdf:
        _render_summary_page(estimate, pdf)
        _render_recommendation_page(estimate, pdf)

    return str(path.resolve())


def _render_summary_page(estimate: QuotaEstimate, pdf: PdfPages) -> None:
    fig, ax = plt.subplots(figsize=(11, 8.5))
    ax.axis("off")
    ax.text(0.05, 0.92, f"Quota Estimate - {estimate.customer_name}", fontsize=20, weight="bold")
    ax.text(0.05, 0.86, f"Generated {estimate.generated_at.isoformat(timespec='seconds')}", fontsize=10)
    ax.text(0.05, 0.76, f"Trailing revenue: {_currency(estimate.trailing_revenue_total)}", fontsize=14)
    ax.text(0.05, 0.70, f"Recommended quota: {_currency(estimate.recommended_quota_total)}", fontsize=14)
    ax.text(0.05, 0.64, f"Overall growth: {_percent(estimate.overall_growth_rate)}", fontsize=14)
    ax.text(0.05, 0.54, "Market context", fontsize=12, weight="bold")
    ax.text(0.05, 0.49, _wrap_text(estimate.research_context.summary, 110), fontsize=10, va="top")
    ax.text(0.05, 0.30, "Methodology", fontsize=12, weight="bold")
    ax.text(0.05, 0.25, _wrap_text(estimate.methodology, 110), fontsize=10, va="top")
    pdf.savefig(fig, bbox_inches="tight")  # type: ignore[no-untyped-call]
    plt.close(fig)


def _render_recommendation_page(estimate: QuotaEstimate, pdf: PdfPages) -> None:
    categories = [f"{item.territory}\n{item.category}" for item in estimate.recommendations]
    trailing = [item.trailing_revenue for item in estimate.recommendations]
    projected = [item.recommended_quota for item in estimate.recommendations]

    fig, ax = plt.subplots(figsize=(11, 8.5))
    x_positions = range(len(categories))
    width = 0.35
    ax.bar([index - width / 2 for index in x_positions], trailing, width, label="Trailing Revenue", color="#0078D4")
    ax.bar([index + width / 2 for index in x_positions], projected, width, label="Recommended Quota", color="#50E6FF")
    ax.set_title("Quota Recommendations by Territory and Category")
    ax.set_ylabel("Revenue")
    ax.set_xticks(list(x_positions))
    ax.set_xticklabels(categories, rotation=30, ha="right", fontsize=8)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _: f"${value / 1000:,.0f}K"))
    ax.legend()
    fig.tight_layout()
    pdf.savefig(fig, bbox_inches="tight")  # type: ignore[no-untyped-call]
    plt.close(fig)


def _append_header(sheet: Any, headers: list[str]) -> None:
    append = getattr(sheet, "append")
    append(headers)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autosize_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
        sheet.column_dimensions[column_letter].width = max(12, max_length + 2)


def _currency(value: float) -> str:
    return f"${value:,.0f}"


def _percent(value: float) -> str:
    return f"{value * 100:.1f}%"


def _wrap_text(value: str, width: int) -> str:
    words = value.split()
    lines: list[str] = []
    current = ""
    for word in words:
        if len(current) + len(word) + 1 > width:
            lines.append(current)
            current = word
        else:
            current = f"{current} {word}".strip()
    if current:
        lines.append(current)
    return "\n".join(lines)
