"""Unit tests for the quota estimation pipeline and artifact renderers."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

from src.agents.quota_estimator.pipeline import (
    build_quota_estimate,
    demo_sales_rows,
    demo_workiq_activity,
    generate_quota_estimation_report,
)

_FIXED_AS_OF = date(2026, 6, 1)
_FIXED_GENERATED_AT = datetime(2026, 6, 1, 9, 0, 0)


def _sales_rows() -> list[dict[str, object]]:
    return [
        {
            "territory": "Northwest",
            "category": "Novelty Items",
            "order_date": "2025-07-01",
            "revenue": 100000,
            "quantity": 200,
        },
        {
            "territory": "Northwest",
            "category": "Novelty Items",
            "order_date": "2026-02-01",
            "revenue": 130000,
            "quantity": 240,
        },
        {
            "territory": "Southwest",
            "category": "Toys",
            "order_date": "2025-08-15",
            "revenue": 75000,
            "quantity": 180,
        },
        {
            "territory": "Southwest",
            "category": "Toys",
            "order_date": "2026-03-15",
            "revenue": 90000,
            "quantity": 210,
        },
    ]


def _research_data() -> dict[str, object]:
    return {
        "summary": "Novelty goods demand is expanding with 12% year-over-year retail growth.",
        "key_metrics": {"revenue_yoy_growth": "12%"},
        "articles": [
            {
                "title": "Retail novelty goods expand",
                "source": "Example Research",
                "url": "https://example.com/research",
            }
        ],
    }


def _workiq_activity(score: str = "High") -> dict[str, object]:
    return {
        "source": "mock WorkIQ",
        "engagement_score": score,
        "last_contact": "2026-05-28",
        "recent_activity": [
            {"type": "email", "subject": "FY27 planning"},
            {"type": "meeting", "subject": "Quota workshop"},
        ],
    }


def test_build_quota_estimate_groups_sales_and_includes_context() -> None:
    estimate = build_quota_estimate(
        customer_name="Tailspin Toys",
        sales_rows=_sales_rows(),
        research_data=_research_data(),
        workiq_activity=_workiq_activity(),
    )

    assert estimate.customer_name == "Tailspin Toys"
    assert len(estimate.recommendations) == 2
    assert estimate.trailing_revenue_total == 395000
    assert estimate.recommended_quota_total > estimate.trailing_revenue_total
    assert estimate.research_context.growth_rate_hint == pytest.approx(0.12)
    assert estimate.workiq_activity.activity_count == 2
    assert any("SalesOrderHeader" in citation for citation in estimate.citations)
    assert any("Retail novelty goods expand" in citation for citation in estimate.citations)


def test_build_quota_estimate_accepts_databricks_genie_rows() -> None:
    estimate = build_quota_estimate(
        customer_name="Tailspin Toys",
        sales_rows=[
            {
                "sales_territory": "Northwest",
                "productCategory": "Novelty Items",
                "orderDate": "2025-07-01",
                "net_sales_amount": "100,000",
                "units_sold": 200,
                "source_platform": "databricks",
            },
            {
                "sales_territory": "Northwest",
                "productCategory": "Novelty Items",
                "orderDate": "2026-02-01",
                "net_sales_amount": "130,000",
                "units_sold": 240,
                "source_platform": "databricks",
            },
        ],
    )

    assert estimate.trailing_revenue_total == 230000
    assert "Databricks Genie" in estimate.methodology
    assert any("Unity Catalog" in citation for citation in estimate.citations)


def test_build_quota_estimate_data_source_override_controls_citation() -> None:
    estimate = build_quota_estimate(
        customer_name="Tailspin Toys",
        sales_rows=_sales_rows(),
        data_source="databricks",
    )

    assert "Databricks Genie" in estimate.methodology
    assert estimate.citations[0] == "Databricks Genie query over Unity Catalog sales tables."


def test_unknown_data_source_is_rejected() -> None:
    with pytest.raises(ValueError, match="Unsupported sales data source"):
        build_quota_estimate(
            customer_name="Tailspin Toys",
            sales_rows=_sales_rows(),
            data_source="spreadsheet",
        )


def test_empty_sales_rows_error_is_platform_neutral() -> None:
    with pytest.raises(ValueError, match="at least one sales row") as exc_info:
        build_quota_estimate(customer_name="Tailspin Toys", sales_rows=[])

    assert "Fabric" not in str(exc_info.value)


def test_workiq_engagement_adjusts_recommended_quota() -> None:
    high = build_quota_estimate(
        customer_name="Tailspin Toys",
        sales_rows=_sales_rows(),
        research_data=_research_data(),
        workiq_activity=_workiq_activity("High"),
    )
    low = build_quota_estimate(
        customer_name="Tailspin Toys",
        sales_rows=_sales_rows(),
        research_data=_research_data(),
        workiq_activity=_workiq_activity("Low"),
    )

    assert high.recommended_quota_total > low.recommended_quota_total


def test_generate_quota_estimation_report_creates_artifacts(tmp_path: Path) -> None:
    result = generate_quota_estimation_report(
        customer_name="Tailspin Toys",
        sales_rows=_sales_rows(),
        research_data=_research_data(),
        workiq_activity=_workiq_activity(),
        output_dir=tmp_path,
        generated_at=_FIXED_GENERATED_AT,
    )

    assert result["scenario"] == "base"
    artifacts = result["artifacts"]
    assert isinstance(artifacts, dict)
    assert set(artifacts) == {"xlsx", "html", "pdf"}

    workbook = openpyxl.load_workbook(artifacts["xlsx"])
    assert workbook.sheetnames == ["Summary", "Recommendations", "Sales Detail", "Methodology", "Assumptions"]
    assert workbook["Assumptions"].cell(row=1, column=6).value == "Scenario Adjustment"

    html_text = Path(str(artifacts["html"])).read_text(encoding="utf-8")
    assert "<html" in html_text.lower()
    assert "data:image/png;base64," in html_text

    pdf_path = Path(str(artifacts["pdf"]))
    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0


def test_scenarios_order_quota_conservative_base_aggressive() -> None:
    totals = {}
    for scenario in ("conservative", "base", "aggressive"):
        estimate = build_quota_estimate(
            customer_name="Tailspin Toys",
            sales_rows=_sales_rows(),
            research_data=_research_data(),
            workiq_activity=_workiq_activity(),
            scenario=scenario,
        )
        assert estimate.scenario == scenario
        totals[scenario] = estimate.recommended_quota_total

    assert totals["conservative"] < totals["base"] < totals["aggressive"]


def test_unknown_scenario_is_rejected() -> None:
    with pytest.raises(ValueError, match="Unsupported scenario"):
        build_quota_estimate(
            customer_name="Tailspin Toys",
            sales_rows=_sales_rows(),
            scenario="moonshot",
        )


def test_synthetic_workiq_fallback_drives_engagement_signal() -> None:
    """The synthetic WorkIQ fallback works without real credentials and lifts engagement."""
    activity = demo_workiq_activity("Tailspin Toys", as_of=_FIXED_AS_OF)
    assert "synthetic" in str(activity["source"]).lower()

    estimate = build_quota_estimate(
        customer_name="Tailspin Toys",
        sales_rows=demo_sales_rows(as_of=_FIXED_AS_OF),
        workiq_activity=activity,
        generated_at=_FIXED_GENERATED_AT,
    )

    assert estimate.workiq_activity.activity_count == 4
    assert estimate.recommendations[0].engagement_adjustment > 0


def test_demo_dates_are_relative_to_as_of() -> None:
    rows = demo_sales_rows(as_of=_FIXED_AS_OF)
    order_dates = sorted(str(row["order_date"]) for row in rows)
    assert order_dates[-1] <= _FIXED_AS_OF.isoformat()
    assert order_dates[0] < order_dates[-1]


def test_generate_quota_estimation_report_rejects_unsupported_format(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="Unsupported quota report format"):
        generate_quota_estimation_report(
            customer_name="Tailspin Toys",
            sales_rows=_sales_rows(),
            output_dir=tmp_path,
            formats=["xlsx", "docx"],
        )


def test_generate_quota_estimation_report_records_databricks_methodology(tmp_path: Path) -> None:
    result = generate_quota_estimation_report(
        customer_name="Tailspin Toys",
        sales_rows=[
            {
                "salesTerritory": "Northwest",
                "productCategory": "Toys",
                "orderDate": "2025-07-01",
                "salesAmount": 100000,
                "quantitySold": 200,
            },
            {
                "salesTerritory": "Northwest",
                "productCategory": "Toys",
                "orderDate": "2026-02-01",
                "salesAmount": 120000,
                "quantitySold": 250,
            },
        ],
        data_source="databricks",
        output_dir=tmp_path,
        formats=["html"],
        generated_at=_FIXED_GENERATED_AT,
    )

    assert "Databricks Genie" in str(result["methodology"])
    assert any("Unity Catalog" in citation for citation in result["citations"])


def test_build_quota_estimate_requires_sales_rows() -> None:
    with pytest.raises(ValueError, match="sales_rows must include"):
        build_quota_estimate(customer_name="Tailspin Toys", sales_rows=[])


def test_end_to_end_demo_artifacts_smoke(tmp_path: Path) -> None:
    """CI smoke test: the full demo data path produces valid, non-empty artifacts."""
    result = generate_quota_estimation_report(
        customer_name="Wide World Importers",
        sales_rows=demo_sales_rows(as_of=_FIXED_AS_OF),
        workiq_activity=demo_workiq_activity("Wide World Importers", as_of=_FIXED_AS_OF),
        scenario="aggressive",
        output_dir=tmp_path,
        generated_at=_FIXED_GENERATED_AT,
    )

    artifacts = result["artifacts"]
    assert isinstance(artifacts, dict)

    xlsx_path = Path(str(artifacts["xlsx"]))
    assert "aggressive" in xlsx_path.name
    workbook = openpyxl.load_workbook(xlsx_path)
    assert "Assumptions" in workbook.sheetnames
    assert len(workbook.sheetnames) == 5

    html_text = Path(str(artifacts["html"])).read_text(encoding="utf-8")
    assert html_text.count("data:image/png;base64,") == 1
    assert "</html>" in html_text.lower()

    assert Path(str(artifacts["pdf"])).stat().st_size > 1000
